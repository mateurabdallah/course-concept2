"""Parsing des fiches d'engagement Excel et fusion en listes de départ par catégorie."""

import io
import json
import os

import openpyxl

DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "startlist_data.json")


def parse_engagement_file(file_bytes, filename_hint=""):
    """Parse une fiche d'engagement et retourne (nom_club, liste d'entrées)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    club_name = None
    for row in rows[:6]:
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower().startswith("club"):
                parts = cell.split(":", 1)
                if len(parts) > 1 and parts[1].strip():
                    club_name = parts[1].strip()
                break
        if club_name:
            break
    if not club_name:
        club_name = filename_hint or "نادي غير محدد"

    categories = {}
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and len(row) > 1 and isinstance(row[1], str) and row[1].strip().lower().startswith("nom"):
            header_row_idx = i
            for j, cell in enumerate(row):
                if j >= 5 and isinstance(cell, str) and cell.strip():
                    categories[j] = cell.strip()
            break

    entries = []
    if header_row_idx is None:
        return club_name, entries

    for row in rows[header_row_idx + 1:]:
        if not row or row[0] is None:
            continue
        try:
            float(row[0])
        except (TypeError, ValueError):
            continue
        nom = (row[1] or "").strip() if isinstance(row[1], str) else ""
        prenom = (row[2] or "").strip() if isinstance(row[2], str) else ""
        full_name = f"{prenom} {nom}".strip()
        if not full_name:
            continue
        for col_idx, cat in categories.items():
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip().lower()
                if val == "x":
                    entries.append({"participant": full_name, "club": club_name, "category": cat})

    return club_name, entries


def load_startlist_data():
    if not os.path.exists(DATA_PATH):
        return {"files": [], "entries": []}
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_startlist_data(data):
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def add_file_entries(filename, club_name, entries):
    data = load_startlist_data()
    data["files"] = [f for f in data["files"] if f != filename]
    data["entries"] = [e for e in data["entries"] if e.get("_source") != filename]
    for e in entries:
        e["_source"] = filename
    data["files"].append(filename)
    data["entries"].extend(entries)
    save_startlist_data(data)
    return data


def clear_startlist_data():
    save_startlist_data({"files": [], "entries": []})


def group_by_category(entries):
    by_cat = {}
    for e in entries:
        by_cat.setdefault(e["category"], []).append(e)
    return dict(sorted(by_cat.items()))
